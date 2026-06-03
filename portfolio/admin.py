import urllib.request
from io import BytesIO
from django.contrib import admin, messages
from django.urls import path
from django.shortcuts import render, redirect
from django.core.files.base import ContentFile
from django.http import HttpResponse
from django.utils.html import format_html
import openpyxl
from .models import GalleryItem, ContactMessage

# Register ContactMessage Admin
@admin.register(ContactMessage)
class ContactMessageAdmin(admin.ModelAdmin):
    list_display = ('name', 'email', 'phone', 'is_read', 'created_at')
    list_filter = ('is_read', 'created_at')
    search_fields = ('name', 'email', 'phone', 'message')
    readonly_fields = ('name', 'email', 'phone', 'message', 'created_at')
    actions = ['mark_as_read', 'mark_as_unread', 'export_to_excel']

    def mark_as_read(self, request, queryset):
        queryset.update(is_read=True)
        self.message_user(request, "Selected messages marked as read.", messages.SUCCESS)
    mark_as_read.short_description = "Mark selected messages as read"

    def mark_as_unread(self, request, queryset):
        queryset.update(is_read=False)
        self.message_user(request, "Selected messages marked as unread.", messages.SUCCESS)
    mark_as_unread.short_description = "Mark selected messages as unread"

    def export_to_excel(self, request, queryset):
        import openpyxl
        from django.http import HttpResponse
        
        # Create a new workbook and sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contact Messages"
        
        # Add headers
        headers = ["Name", "Email", "Phone", "Message", "Status", "Date Submitted"]
        ws.append(headers)
        
        from django.utils import timezone
        
        # Add message rows
        for msg in queryset:
            status = "Read" if msg.is_read else "Unread"
            if msg.created_at:
                local_time = timezone.localtime(msg.created_at)
                date_str = local_time.strftime("%Y-%m-%d %H:%M")
            else:
                date_str = "-"
            ws.append([msg.name, msg.email, msg.phone, msg.message, status, date_str])
            
        # Prepare HTTP Response with Excel mime-type
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=contact_messages.xlsx"
        wb.save(response)
        return response
    export_to_excel.short_description = "Export selected messages to Excel"


# Register GalleryItem Admin with custom Excel Importer
@admin.register(GalleryItem)
class GalleryItemAdmin(admin.ModelAdmin):
    list_display = ('thumbnail', 'title', 'category', 'is_visible', 'order', 'created_at')
    list_editable = ('title', 'category', 'is_visible', 'order')
    list_filter = ('category', 'is_visible', 'created_at')
    search_fields = ('title', 'description')
    ordering = ('order', '-created_at')

    def thumbnail(self, obj):
        if obj.image:
            return format_html('<img src="{}" style="width: 45px; height: 45px; object-fit: cover; border-radius: 4px;" />', obj.image.url)
        return "-"
    thumbnail.short_description = 'Preview'

    # 1. Inject custom URL routes into Admin
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-excel/', self.admin_site.admin_view(self.import_excel_view), name='portfolio_galleryitem_import_excel'),
            path('download-template/', self.admin_site.admin_view(self.download_template_view), name='portfolio_galleryitem_download_template'),
        ]
        return custom_urls + urls

    # 2. View to process Excel Uploads
    def import_excel_view(self, request):
        if request.method == 'POST' and request.FILES.get('excel_file'):
            excel_file = request.FILES['excel_file']
            
            try:
                # Open Excel Workbook
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.active
                
                success_count = 0
                error_count = 0
                error_details = []

                # Columns expected: Title, Category, Image URL, Description, Order
                # Let's read from row 2 (skipping header)
                for r_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    # Guard: stop if empty row
                    if not any(row):
                        continue
                        
                    title = row[0]
                    category_input = str(row[1] or '').strip()
                    image_url = row[2]
                    description = row[3]
                    order = row[4]

                    # Category mapping/normalization
                    category_mapping = {
                        'elevation design': 'Elevation Design',
                        'residential houses': 'Residential Houses',
                        'residential villas': 'Residential Villas',
                        'interior design': 'Interior Design',
                        'landscape design': 'Landscape Design',
                        'commercial buildings': 'Commercial Buildings',
                        'renovation & remodeling': 'Renovation & Remodeling',
                        'renovation and remodeling': 'Renovation & Remodeling',
                        # Fallbacks for old categories
                        'villas': 'Residential Villas',
                        'interiors': 'Interior Design',
                        'commercial': 'Commercial Buildings',
                        'renovation': 'Renovation & Remodeling',
                    }

                    category_lower = category_input.lower()
                    if not category_input or category_lower not in category_mapping:
                        error_count += 1
                        error_details.append(
                            f"Row {r_idx}: Invalid category '{category_input}'. "
                            "Must be one of the seven new categories (e.g. 'Elevation Design', 'Residential Houses', etc.)."
                        )
                        continue
                    
                    category = category_mapping[category_lower]
                    
                    if not image_url:
                        error_count += 1
                        error_details.append(f"Row {r_idx}: Missing 'Image URL'. Image is required.")
                        continue

                    # Attempt to download image
                    try:
                        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
                        img_req = urllib.request.Request(image_url, headers=headers)
                        with urllib.request.urlopen(img_req, timeout=10) as response:
                            img_data = response.read()
                            
                        # Extract filename from URL
                        filename = image_url.split('/')[-1].split('?')[0]
                        if not filename or '.' not in filename:
                            filename = f"imported_image_{r_idx}.jpg"
                            
                        img_file = ContentFile(img_data, name=filename)
                    except Exception as img_err:
                        error_count += 1
                        error_details.append(f"Row {r_idx}: Failed to download image from '{image_url}'. Error: {str(img_err)}")
                        continue

                    # Save record to database
                    GalleryItem.objects.create(
                        title=title or '',
                        category=category,
                        image=img_file,
                        description=description or '',
                        is_visible=True,
                        order=int(order or 0)
                    )
                    success_count += 1

                # Display detailed import reports
                if success_count > 0:
                    messages.success(request, f"Successfully imported {success_count} gallery items from Excel!")
                if error_count > 0:
                    messages.error(request, f"Failed to import {error_count} rows. Errors:\n" + "\n".join(error_details))

                return redirect('admin:portfolio_galleryitem_changelist')

            except Exception as e:
                messages.error(request, f"Failed to read Excel file: {str(e)}")
                return redirect('admin:portfolio_galleryitem_changelist')

        # If GET, render upload form template
        context = {
            **self.admin_site.each_context(request),
            'title': 'Import Gallery Items from Excel',
            'opts': self.model._meta,
        }
        return render(request, 'admin/portfolio/galleryitem/import_excel.html', context)

    # 3. View to download simple template
    def download_template_view(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Gallery Import Template"

        # Headers
        ws.append(["Title", "Category", "Image URL", "Description", "Order"])
        
        # Sample rows for guide
        ws.append(["Luxury Villa Beachfront", "Residential Villas", "https://images.unsplash.com/photo-1564013799919-ab600027ffc6?w=600&q=80", "Modern beachfront villa with pool", 1])
        ws.append(["Elegant Penthouse Living", "Interior Design", "https://images.unsplash.com/photo-1600566753190-17f0baa2a6c3?w=600&q=80", "Minimalist living room interior design", 2])

        # Write buffer
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=gallery_import_template.xlsx"
        wb.save(response)
        return response
